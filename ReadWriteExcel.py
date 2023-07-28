import openpyxl
import datetime


class OperExcel():

    # 创建一个新的excel并写入数据
    def Create_excel(Path, filename):
        # 创建一个 workbook（Excel）
        workbook = openpyxl.Workbook()
        # Workbook实例化后会自带一个名为sheet的页，workbook.active调用被激活的 worksheet
        worksheet = workbook.active
        worksheet.title = "原始表单"
        # 设置单元格内容
        worksheet['A1'] = 42
        # 设置一行内容
        # append完后追加内容
        worksheet.append([1, 2, 3])
        worksheet.append([1, 2, 3, 4, 5])
        # python 数据类型可以被自动转换
        worksheet['A2'] = datetime.datetime.now()
        # 保存 Excel 文件
        workbook.save(Path + filename)
        print(Path + filename, "创建完成！")

    # 操作sheet页 ，并创建新表单，重命名表单
    def Data_Read(filename):
        workbook = openpyxl.load_workbook(filename)
        # 创建一个新的sheet页
        if "新表单" not in workbook.sheetnames:
            sht1 = workbook.create_sheet('新表单', 1)
            sht1.append(["a", "b", "c ddd", "eee", "创建新的表单"])
        else:
            pass
        # 复制"新表单"页
        sht2_copy = workbook.copy_worksheet(workbook.worksheets[1])
        sht2_copy.title = "复制表单"
        print("新表单复制完成--")
        sht3 = workbook.copy_worksheet(sht2_copy)
        sht3.title = "复制表单2"
        print("复制表单二完成")
        # del根据sheet名字删除，不能使用索引号
        del workbook["复制表单"]
        print("已删除复制表单")

        del workbook["复制表单2"]
        # 保存
        workbook.save(filename)
        # 显示文档中包含的 表单 名称

    # 操作单元格
    def Oper_Cell(filename, sheetname):
        # 加载excel
        workbook = openpyxl.load_workbook(filename)
        sht1 = workbook.create_sheet("操作单元格")  # 创建一个sheet
        # 通过单元格名称设置
        sht1["A1"] = 3.1415926
        sht1["B2"] = "你好！！！！！！！"
        # 通过行列坐标设置
        pos = sht1.cell(row=4, column=2, value=10)
        # 批量操作单元格
        # 列循环
        ws2 = workbook.create_sheet("列循环输出")
        for cell in ws2["A"]:
            print(cell.value)
        # 行循环
        for cell in ws2["1"]:
            print(cell.value)
        # 操作多列循环
        ws3 = workbook.create_sheet("多列多行循环")
        for column in ws3['A:C']:
            for cell in column:
                print(cell.value)
        # 操作多行循环
        for row in ws3['1:3']:
            for cell in row:
                print(cell.value)
        # 指定范围
        for row in ws3['A1:C3']:
            for cell in row:
                print(cell.value)

    # def FontSet(filename, sheetname):
    #
    # def Merge(filename, sheetname):


if __name__ == '__main__':
    # 如果文档路径/被转义可以用\代替。
    # OperExcel.Create_excel("D:/Oracle", "test.xlsx")
    OperExcel.Data_Read("D:/Oracle/test.xlsx")
    # OperExcel.Oper_Cell()
