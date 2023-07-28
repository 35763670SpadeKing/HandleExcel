from copy import copy
import openpyxl
import Transform
import os

# 转换xls到xlsx文件格式
# xls_folder = r'C:\xls'
# xlsx_folder = r'C:\xlsx'
# Transform.batch_convert(xls_folder, xlsx_folder)

# 定义批量处理的输入输出文件夹
before_folder = "D:\离线数据\conv_xlsx\\"
after_floder = "C:\\Users\\Mango\\Desktop\\after\\"

# 循环依次获取以xlsx结尾的文件，并按格式处理
filename = [f for f in os.listdir(before_folder) if f.endswith('.xlsx')]
print(filename)
for f in filename:
    # 打开文件
    workbook = openpyxl.load_workbook(before_folder + f)
    worksheet = workbook.active
    print(before_folder + f)
    # 获取表的行数
    last_row = worksheet.max_row
    if last_row < 7:
        # 如果行数小于 7 空表，退出循环
        continue  # 不处理当前文件，跳到下一个文件。
        # 如果是break 则跳出当前循环不处理所有文件。
    # 删除表头1-6行
    # worksheet.delete_rows(1, 6)
    # 遍历工作表中的每一行
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        # 先修改表头拓展当前表格字段
        worksheet['U6'].value = worksheet['B2'].value
        worksheet['V6'].value = worksheet['Q2'].value
        worksheet['W6'].value = worksheet['B3'].value
        worksheet['X6'].value = worksheet['B5'].value
        # 填充拓展的字段值，为表头的固定值
        # 获取 各个固定单元格的值
        u_value = worksheet['C2'].value
        v_value = worksheet['R2'].value
        w_value = worksheet['C3'].value
        x_value = worksheet['C5'].value

    # 遍历表格新增的列并且填充
    for row in range(7, last_row + 1):
        worksheet.cell(row=row, column=21, value=u_value)
        worksheet.cell(row=row, column=22, value=v_value)
        worksheet.cell(row=row, column=23, value=w_value)
        worksheet.cell(row=row, column=24, value=x_value)

    # 删除表格中的统计行，B列以卡号开头。从后往前删除
    for row in reversed(list(worksheet.iter_rows(min_row=6))):
        if str(row[1].value).startswith("卡号"):
            # 拆分合并的单元格
            worksheet.unmerge_cells(start_row=row[0].row, start_column=4, end_row=row[0].row, end_column=16)
            worksheet.unmerge_cells(start_row=row[0].row, start_column=17, end_row=row[0].row, end_column=19)
            # 合并成目标格式
            worksheet.merge_cells(start_row=row[0].row, start_column=6, end_row=row[0].row, end_column=7)
            worksheet.merge_cells(start_row=row[0].row, start_column=8, end_row=row[0].row, end_column=9)
            worksheet.merge_cells(start_row=row[0].row, start_column=11, end_row=row[0].row, end_column=12)
            worksheet.merge_cells(start_row=row[0].row, start_column=13, end_row=row[0].row, end_column=14)
            worksheet.merge_cells(start_row=row[0].row, start_column=15, end_row=row[0].row, end_column=16)
            worksheet.merge_cells(start_row=row[0].row, start_column=17, end_row=row[0].row, end_column=18)
            # print(row[0].row)
            worksheet.delete_rows(row[0].row)

    last_row = worksheet.max_row
    # 删除表尾1行
    worksheet.delete_rows(last_row)
    # 将更改保存到工作簿中
    workbook.save(after_floder + f)

