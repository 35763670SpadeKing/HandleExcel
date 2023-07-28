import openpyxl
import os

# 要合并的文件夹路径
folder_path = r"D:\离线数据\xlsx"

# 新建一个工作簿
wb_merged = openpyxl.Workbook()
ws_merged = wb_merged.active

# 记录表头，只需要保留第一个表头
header = None

# 遍历文件夹中的所有 xlsx 文件
for root, dirs, files in os.walk(folder_path):
    for file in files:
        if file.endswith(".xlsx"):
            # 打开 xlsx 文件
            wb = openpyxl.load_workbook(os.path.join(root, file))
            ws = wb.active

            # 如果是第一个文件，记录表头
            if header is None:
                header = []
                # 获取表头
                for row in ws.iter_rows(min_row=1, max_row=1):
                    header.append([cell.value for cell in row])
                # 将表头写入合并文件
                for row in header:
                    ws_merged.append(row)

            # 获取数据
            for row in ws.iter_rows(min_row=2):
                values = [cell.value for cell in row]
                # for cell in row:
                #      if cell.value == '1000113300006170354':
                #        print(file)
                ws_merged.append(values)

# 保存合并文件
wb_merged.save(os.path.join(folder_path, "merged.xlsx"))